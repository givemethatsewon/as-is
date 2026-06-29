from __future__ import annotations

import csv
from decimal import Decimal
from io import BytesIO, StringIO
from typing import Any

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import func, select
from sqlalchemy.orm import Session, aliased

from app.models import ExportAllocation, ExportRequirement, ImportLot, UploadBatch
from app.services.summaries import dashboard_summary


STATUS_LABELS = {
    "available": "사용 가능",
    "expiring_soon": "만료 예정",
    "expired": "기한 초과",
    "used_up": "소진",
    "blocked": "사용 보류",
    "pending": "매칭 대기",
    "matched": "매칭 완료",
    "partial_matched": "일부 매칭",
    "insufficient_stock": "재고 부족",
}

SHEET_CONFIG = {
    "수입신고별 잔량": {
        "source": "import_lots_with_remaining",
        "headers": {
            "import_declaration_no": "수입신고번호",
            "import_accepted_date": "수리일",
            "origin": "원산지",
            "hs_code": "HS 코드",
            "line_no": "란번호",
            "row_no": "행번호",
            "part_number": "품번",
            "spec": "규격/품명",
            "import_qty": "수입 수량",
            "used_qty": "사용 수량",
            "remaining_qty": "남은 수량",
            "qty_unit": "수량 단위",
            "status": "상태",
        },
    },
    "수출건별 수입근거 매칭": {
        "source": "export_match_allocations",
        "headers": {
            "export_date": "수출일",
            "order_no": "Order No",
            "seq_no": "Seq No",
            "part_number": "품번",
            "description": "설명",
            "unit_price": "단가",
            "required_qty": "필요 수량",
            "amount": "금액",
            "matched_qty": "매칭 수량",
            "import_declaration_no": "수입신고번호",
            "import_accepted_date": "수입수리일",
            "origin": "원산지",
            "hs_code": "HS 코드",
            "line_no": "수입 란번호",
            "row_no": "수입 행번호",
            "remaining_qty_after": "매칭 후 잔량",
            "shortage_qty": "부족 수량",
            "match_status": "매칭 상태",
            "hs_code_warning": "HS 코드 확인",
            "expected_refund_amount": "환급예상",
        },
    },
    "요약": {
        "source": "dashboard_summary",
        "headers": {
            "metric": "항목",
            "value": "값",
        },
    },
    "품번별 재고 요약": {
        "source": "inventory_summary",
        "headers": {
            "part_number": "품번",
            "origin": "원산지",
            "total_imported_qty": "총 수입 수량",
            "total_exported_qty": "총 사용 수량",
            "remaining_qty": "남은 수량",
        },
    },
}


def import_lot_rows(db: Session) -> list[dict[str, Any]]:
    lots = db.scalars(
        select(ImportLot)
        .outerjoin(UploadBatch, ImportLot.upload_batch_id == UploadBatch.id)
        .where((ImportLot.upload_batch_id.is_(None)) | (UploadBatch.invalidated_at.is_(None)))
        .order_by(ImportLot.part_number, ImportLot.origin, ImportLot.import_accepted_date)
    )
    return [
        {
            "import_declaration_no": lot.import_declaration_no,
            "import_accepted_date": lot.import_accepted_date.isoformat(),
            "origin": lot.origin,
            "hs_code": lot.hs_code,
            "line_no": lot.line_no,
            "row_no": lot.row_no,
            "part_number": lot.part_number,
            "spec": lot.spec,
            "import_qty": lot.import_qty,
            "used_qty": lot.used_qty,
            "remaining_qty": lot.remaining_qty,
            "qty_unit": lot.qty_unit,
            "status": STATUS_LABELS.get(lot.status, lot.status),
        }
        for lot in lots
    ]


def allocation_rows(db: Session) -> list[dict[str, Any]]:
    import_batch = aliased(UploadBatch)
    export_batch = aliased(UploadBatch)
    allocation_records = db.execute(
        select(ExportRequirement, ExportAllocation, ImportLot)
        .join(ExportAllocation, ExportAllocation.export_requirement_id == ExportRequirement.id)
        .join(ImportLot, ExportAllocation.import_lot_id == ImportLot.id)
        .outerjoin(import_batch, ImportLot.upload_batch_id == import_batch.id)
        .outerjoin(export_batch, ExportRequirement.upload_batch_id == export_batch.id)
        .where((ImportLot.upload_batch_id.is_(None)) | (import_batch.invalidated_at.is_(None)))
        .where((ExportRequirement.upload_batch_id.is_(None)) | (export_batch.invalidated_at.is_(None)))
        .order_by(ExportRequirement.export_date, ExportRequirement.part_number, ImportLot.import_accepted_date)
    ).all()
    allocations_by_export: dict[str, list[tuple[ExportAllocation, ImportLot]]] = {}
    for export, allocation, lot in allocation_records:
        allocations_by_export.setdefault(export.id, []).append((allocation, lot))

    active_exports = db.scalars(
        select(ExportRequirement)
        .outerjoin(UploadBatch, ExportRequirement.upload_batch_id == UploadBatch.id)
        .where((ExportRequirement.upload_batch_id.is_(None)) | (UploadBatch.invalidated_at.is_(None)))
        .order_by(ExportRequirement.export_date, ExportRequirement.part_number, ExportRequirement.id)
    ).all()

    rows: list[dict[str, Any]] = []
    for export in active_exports:
        matched_qty = 0
        for allocation, lot in allocations_by_export.get(export.id, []):
            matched_qty += allocation.matched_qty
            rows.append(_allocation_report_row(export, allocation, lot))

        shortage_qty = max(export.required_qty - matched_qty, 0)
        if shortage_qty > 0 and export.status in {"partial_matched", "insufficient_stock"}:
            rows.append(_no_match_report_row(export, shortage_qty))

    return rows


def _base_export_report_row(export: ExportRequirement) -> dict[str, Any]:
    return {
        "export_date": export.export_date.isoformat(),
        "order_no": export.order_no,
        "seq_no": export.seq_no,
        "part_number": export.part_number,
        "description": export.description,
        "unit_price": export.unit_price,
        "required_qty": export.required_qty,
        "amount": export.amount,
    }


def _allocation_report_row(export: ExportRequirement, allocation: ExportAllocation, lot: ImportLot) -> dict[str, Any]:
    row = _base_export_report_row(export)
    row.update(
        {
            "matched_qty": allocation.matched_qty,
            "import_declaration_no": lot.import_declaration_no,
            "import_accepted_date": lot.import_accepted_date.isoformat(),
            "origin": lot.origin,
            "hs_code": lot.hs_code,
            "line_no": lot.line_no,
            "row_no": lot.row_no,
            "remaining_qty_after": allocation.remaining_qty_after,
            "shortage_qty": 0,
            "match_status": STATUS_LABELS.get(export.status, export.status),
            "hs_code_warning": allocation.hs_code_warning,
            "expected_refund_amount": allocation.expected_refund_amount,
        }
    )
    return row


def _no_match_report_row(export: ExportRequirement, shortage_qty: int) -> dict[str, Any]:
    row = _base_export_report_row(export)
    row.update(
        {
            "matched_qty": 0,
            "import_declaration_no": "NO MATCH",
            "import_accepted_date": "",
            "origin": export.origin,
            "hs_code": "",
            "line_no": "",
            "row_no": "",
            "remaining_qty_after": "",
            "shortage_qty": shortage_qty,
            "match_status": "NO MATCH",
            "hs_code_warning": "",
            "expected_refund_amount": None,
        }
    )
    return row


def dashboard_rows(db: Session) -> list[dict[str, Any]]:
    summary = dashboard_summary(db)
    return [{"metric": key, "value": value} for key, value in summary.items()]


def inventory_summary_rows(db: Session) -> list[dict[str, Any]]:
    rows = db.execute(
        select(
            ImportLot.part_number,
            ImportLot.origin,
            func.coalesce(func.sum(ImportLot.import_qty), 0),
            func.coalesce(func.sum(ImportLot.used_qty), 0),
            func.coalesce(func.sum(ImportLot.remaining_qty), 0),
        )
        .outerjoin(UploadBatch, ImportLot.upload_batch_id == UploadBatch.id)
        .where((ImportLot.upload_batch_id.is_(None)) | (UploadBatch.invalidated_at.is_(None)))
        .group_by(ImportLot.part_number, ImportLot.origin)
        .order_by(ImportLot.part_number, ImportLot.origin)
    )
    return [
        {
            "part_number": part_number,
            "origin": origin,
            "total_imported_qty": imported,
            "total_exported_qty": exported,
            "remaining_qty": remaining,
        }
        for part_number, origin, imported, exported, remaining in rows
    ]


def rows_to_csv(rows: list[dict[str, Any]]) -> str:
    output = StringIO()
    if not rows:
        return ""
    writer = csv.DictWriter(output, fieldnames=list(rows[0].keys()))
    writer.writeheader()
    writer.writerows(rows)
    return output.getvalue()


def refund_report_xlsx(db: Session) -> bytes:
    return video_style_export_result_xlsx(db)


def video_style_export_result_xlsx(db: Session) -> bytes:
    output = BytesIO()
    headers = {
        "order_no": "Order No",
        "seq_no": "Seq No",
        "part_number": "Part Number",
        "description": "Description",
        "unit_price": "U/Price",
        "required_qty": "Ready to Ship Qty",
        "amount": "Amount",
        "origin": "원산지",
        "import_declaration_no": "수입신고번호",
        "import_accepted_date": "수리일",
        "hs_code": "세번",
        "line_no": "수입 란번호",
        "row_no": "수입 행번호",
        "matched_qty": "매칭 수량",
        "remaining_qty_after": "매칭 후 잔량",
        "shortage_qty": "부족 수량",
        "match_status": "매칭 상태",
    }
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        frame = pd.DataFrame(allocation_rows(db), columns=headers.keys()).rename(columns=headers)
        frame.to_excel(writer, sheet_name="수출 결과", index=False)
        _style_report_sheet(writer.sheets["수출 결과"])
    output.seek(0)
    return output.read()


def contest_example_report_xlsx(db: Session) -> bytes:
    output = BytesIO()
    sheets = {
        "수출 전 확인용 잔량표": (
            import_lot_rows(db),
            {
                "import_declaration_no": "수입신고번호",
                "import_accepted_date": "수리일",
                "origin": "원산지",
                "hs_code": "세번",
                "line_no": "란번호",
                "row_no": "행번호",
                "part_number": "품번",
                "import_qty": "수입 수량",
                "used_qty": "기매칭 수량",
                "remaining_qty": "잔량",
                "status": "상태",
            },
        ),
        "수출 건별 수입근거 자동기재표": (
            allocation_rows(db),
            {
                "export_date": "수출일",
                "order_no": "Order No",
                "seq_no": "Seq No",
                "part_number": "Part Number",
                "description": "Description",
                "required_qty": "Qty",
                "amount": "Amount",
                "import_declaration_no": "수입신고번호",
                "import_accepted_date": "수리일",
                "origin": "원산지",
                "hs_code": "세번",
                "line_no": "란번호",
                "row_no": "행번호",
                "matched_qty": "매칭 수량",
                "remaining_qty_after": "매칭 후 잔량",
                "shortage_qty": "부족 수량",
                "hs_code_warning": "HS 코드 확인",
            },
        ),
    }
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        for sheet_name, (rows, headers) in sheets.items():
            frame = pd.DataFrame(rows, columns=headers.keys()).rename(columns=headers)
            frame.to_excel(writer, sheet_name=sheet_name, index=False)
            _style_report_sheet(writer.sheets[sheet_name])
    output.seek(0)
    return output.read()


def _style_report_sheet(worksheet) -> None:
    header_fill = PatternFill("solid", fgColor="0F5F50")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center")
    body_alignment = Alignment(vertical="top")

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    worksheet.sheet_view.showGridLines = False

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = body_alignment
            if isinstance(cell.value, (int, float, Decimal)):
                cell.number_format = "#,##0"

    for column_cells in worksheet.columns:
        column_letter = get_column_letter(column_cells[0].column)
        max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
        worksheet.column_dimensions[column_letter].width = min(max(max_length + 4, 12), 38)

    worksheet.row_dimensions[1].height = 24
