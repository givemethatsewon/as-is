from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook, load_workbook
from sqlalchemy import select

from app.models import ExportRequirement


def test_upload_confirm_match_and_download_reports(client):
    import_csv = (
        "import_declaration_no,import_accepted_date,origin,hs_code,line_no,row_no,part_number,spec,import_qty,qty_unit\n"
        "A,2025-01-16,CN,8708309000,004,01,MTG011114,STEERING RACK,1256,PC\n"
    )
    export_csv = (
        "export_date,origin,part_number,required_qty,description,unit_price\n"
        "2025-08-16,CN,MTG011114,800,STEERING RACK,3\n"
    )

    import_preview = client.post(
        "/api/imports/preview",
        files={"file": ("imports.csv", import_csv, "text/csv")},
    )
    assert import_preview.status_code == 200
    import_batch_id = import_preview.json()["batch_id"]
    import_confirm = client.post("/api/imports/confirm", data={"batch_id": import_batch_id})
    assert import_confirm.status_code == 200
    assert import_confirm.json()["inserted_count"] == 1

    export_preview = client.post(
        "/api/exports/preview",
        files={"file": ("exports.csv", export_csv, "text/csv")},
    )
    assert export_preview.status_code == 200
    export_batch_id = export_preview.json()["batch_id"]
    export_confirm = client.post("/api/exports/confirm", data={"batch_id": export_batch_id})
    assert export_confirm.status_code == 200
    assert export_confirm.json()["inserted_count"] == 1

    matching = client.post("/api/matching/run")
    assert matching.status_code == 200
    assert matching.json()["matched_count"] == 1
    assert matching.json()["allocation_count"] == 1

    inventory = client.get("/api/inventory", params={"part_number": "MTG011114", "origin": "CN"})
    assert inventory.status_code == 200
    assert inventory.json()["remaining_qty"] == 456

    xlsx_response = client.get("/api/reports/download.xlsx")
    assert xlsx_response.status_code == 200
    assert xlsx_response.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    assert "video_style_export_result.xlsx" in xlsx_response.headers["content-disposition"]
    assert xlsx_response.content.startswith(b"PK")

    workbook = load_workbook(BytesIO(xlsx_response.content))
    assert workbook.sheetnames == ["수출 결과"]
    sheet = workbook["수출 결과"]
    assert sheet["A1"].value == "Order No"
    assert sheet["A1"].font.bold
    assert sheet["A1"].fill.fgColor.rgb == "000F5F50"
    assert sheet.freeze_panes == "A2"
    assert sheet.auto_filter.ref == sheet.dimensions
    headers = [cell.value for cell in sheet[1]]
    assert headers[:8] == ["Order No", "Seq No", "Part Number", "Description", "U/Price", "Ready to Ship Qty", "Amount", "원산지"]
    assert "수입 란번호" in headers
    assert "수입 행번호" in headers
    assert "수입신고번호" in headers
    assert "수리일" in headers
    assert "매칭 수량" in headers
    assert "매칭 후 잔량" in headers
    assert sheet.cell(row=2, column=headers.index("수입신고번호") + 1).value == "A"


def test_import_preview_reads_video_style_xlsm_stock_sheet(client):
    workbook = Workbook()
    notice = workbook.active
    notice.title = "안내"
    notice.append(["이 시트는 설명용입니다."])
    stock = workbook.create_sheet("원상태진행")
    stock.append(["원상태 수출 재고 관리"])
    stock.append(["수입신고번호", "신고일자", "원산지", "세번", "란번호", "행번호", "판매부번", "규격", "수량", "수량단위", "잔량"])
    stock.append(["4397824102396M", "20240724", "CN", "8708940000", "001", "29", "MTG011114", "STEERING RACK", "280", "PC", "205"])
    content = _workbook_bytes(workbook)

    response = client.post(
        "/api/imports/preview",
        files={
            "file": (
                "원상태진행 TEST 매크로.xlsm",
                content,
                "application/vnd.ms-excel.sheet.macroEnabled.12",
            )
        },
    )

    assert response.status_code == 200
    body = response.json()
    assert body["new_count"] == 1
    assert body["column_mapping"]["import_declaration_no"] == "수입신고번호"
    assert body["column_mapping"]["part_number"] == "판매부번"
    assert body["column_mapping"]["import_qty"] == "수량"


def test_export_preview_reads_video_style_xlsm_export_sheet(client):
    workbook = Workbook()
    stock = workbook.active
    stock.title = "원상태진행"
    stock.append(["수입신고번호", "신고일자", "원산지", "세번", "란번호", "행번호", "판매부번", "규격", "수량", "수량단위"])
    stock.append(["A", "20250116", "CN", "8708309000", "001", "01", "MTG011114", "STEERING RACK", "1256", "PC"])
    exports = workbook.create_sheet("수출")
    exports.append(["INVOICE RIDER"])
    exports.append(["Order No", "Seq No", "Part Number", "Description", "U/Price", "Ready to Ship\nQty", "Amount", "원산지"])
    exports.append(["SOTB20056-250507002-01", "1", "MTG011114", "STEERING RACK", "3.5", "800", "2800", "CN"])
    content = _workbook_bytes(workbook)

    response = client.post(
        "/api/exports/preview",
        files={
            "file": (
                "원상태 수출관리 시연.xlsm",
                content,
                "application/vnd.ms-excel.sheet.macroEnabled.12",
            )
        },
    )

    assert response.status_code == 200
    body = response.json()
    assert body["new_count"] == 1
    assert body["column_mapping"]["part_number"] == "Part Number"
    assert body["column_mapping"]["required_qty"] == "Ready to Ship Qty"
    assert body["column_mapping"]["unit_price"] == "U/Price"
    assert body["column_mapping"]["amount"] == "Amount"
    assert body["column_mapping"]["origin"] == "원산지"
    assert body["column_mapping"]["order_no"] == "Order No"
    assert body["column_mapping"]["seq_no"] == "Seq No"


def test_matching_report_includes_no_match_row_for_shortage(client):
    import_csv = (
        "수입신고번호,declaration_date,원산지,HS Code,란번호,행번호,Part Number,규격,수량,수량단위\n"
        "A,2025-01-16,CN,8708309000,004,01,MTG011114,STEERING RACK,10,PC\n"
    )
    export_csv = (
        "export_date,origin,Order No,Seq No,Part Number,Description,Qty,Amount\n"
        "2025-08-16,CN,SOTB20056-250507002-01,1,MTG011114,STEERING RACK,15,45\n"
    )

    import_preview = client.post("/api/imports/preview", files={"file": ("imports.csv", import_csv, "text/csv")})
    assert import_preview.status_code == 200
    assert client.post("/api/imports/confirm", data={"batch_id": import_preview.json()["batch_id"]}).status_code == 200

    export_preview = client.post("/api/exports/preview", files={"file": ("exports.csv", export_csv, "text/csv")})
    assert export_preview.status_code == 200
    assert client.post("/api/exports/confirm", data={"batch_id": export_preview.json()["batch_id"]}).status_code == 200

    matching = client.post("/api/matching/run")
    assert matching.status_code == 200
    assert matching.json()["partial_matched_count"] == 1

    rows = client.get("/api/reports/export-allocations").json()
    assert len(rows) == 2
    assert rows[0]["order_no"] == "SOTB20056-250507002-01"
    assert rows[0]["seq_no"] == "1"
    assert rows[0]["matched_qty"] == 10
    assert rows[0]["remaining_qty_after"] == 0
    assert rows[1]["order_no"] == "SOTB20056-250507002-01"
    assert rows[1]["match_status"] == "NO MATCH"
    assert rows[1]["matched_qty"] == 0
    assert rows[1]["shortage_qty"] == 5
    assert rows[1]["import_declaration_no"] == "NO MATCH"

    result_response = client.get("/api/reports/download.xlsx")
    workbook = load_workbook(BytesIO(result_response.content))
    sheet = workbook["수출 결과"]
    headers = [cell.value for cell in sheet[1]]
    assert "Order No" in headers
    assert "Seq No" in headers
    assert "부족 수량" in headers
    assert sheet.cell(row=3, column=headers.index("수입신고번호") + 1).value == "NO MATCH"
    assert sheet.cell(row=3, column=headers.index("부족 수량") + 1).value == 5


def test_api_can_undo_export_matching(client, db_session):
    import_csv = (
        "import_declaration_no,import_accepted_date,origin,hs_code,line_no,row_no,part_number,spec,import_qty,qty_unit\n"
        "A,2025-01-16,CN,8708309000,004,01,MTG011114,STEERING RACK,20,PC\n"
    )
    export_csv = "export_date,origin,part_number,required_qty\n2025-08-16,CN,MTG011114,8\n"

    import_preview = client.post("/api/imports/preview", files={"file": ("imports.csv", import_csv, "text/csv")})
    client.post("/api/imports/confirm", data={"batch_id": import_preview.json()["batch_id"]})
    export_preview = client.post("/api/exports/preview", files={"file": ("exports.csv", export_csv, "text/csv")})
    client.post("/api/exports/confirm", data={"batch_id": export_preview.json()["batch_id"]})
    client.post("/api/matching/run")

    export = db_session.scalar(select(ExportRequirement))
    assert export is not None
    assert client.get("/api/inventory", params={"part_number": "MTG011114", "origin": "CN"}).json()["remaining_qty"] == 12

    response = client.post(f"/api/exports/{export.id}/matching/undo")

    assert response.status_code == 200
    assert response.json()["undone_allocation_count"] == 1
    assert client.get("/api/inventory", params={"part_number": "MTG011114", "origin": "CN"}).json()["remaining_qty"] == 20


def test_exports_page_shows_per_item_undo_action_after_matching(client):
    import_csv = (
        "import_declaration_no,import_accepted_date,origin,hs_code,line_no,row_no,part_number,spec,import_qty,qty_unit\n"
        "A,2025-01-16,CN,8708309000,004,01,MTG011114,STEERING RACK,20,PC\n"
    )
    export_csv = (
        "export_date,origin,Order No,Seq No,Part Number,Qty\n"
        "2025-08-16,CN,SOTB20056-250507002-01,1,MTG011114,8\n"
    )

    import_preview = client.post("/api/imports/preview", files={"file": ("imports.csv", import_csv, "text/csv")})
    client.post("/api/imports/confirm", data={"batch_id": import_preview.json()["batch_id"]})
    export_preview = client.post("/api/exports/preview", files={"file": ("exports.csv", export_csv, "text/csv")})
    client.post("/api/exports/confirm", data={"batch_id": export_preview.json()["batch_id"]})
    client.post("/api/matching/run")

    response = client.get("/exports")

    assert response.status_code == 200
    assert "SOTB20056-250507002-01" in response.text
    assert "Seq" in response.text
    assert "이 건 되돌리기" in response.text
    assert "되돌리기는 해당 수출 항목 하나만 원복합니다" in response.text


def test_import_preview_supports_column_aliases(client):
    import_csv = (
        "수입신고번호,declaration_date,원산지,HS Code,란번호,행번호,Part Number,규격,수량,수량단위\n"
        "A,2025-01-16,CN,8708309000,004,01,MTG011114,STEERING RACK,1256,PC\n"
    )

    response = client.post(
        "/api/imports/preview",
        files={"file": ("imports.csv", import_csv, "text/csv")},
    )

    assert response.status_code == 200
    body = response.json()
    assert body["new_count"] == 1
    assert body["column_mapping"]["import_accepted_date"] == "declaration_date"
    assert body["column_mapping"]["import_declaration_no"] == "수입신고번호"
    assert body["column_mapping"]["part_number"] == "Part Number"


def _workbook_bytes(workbook: Workbook) -> bytes:
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def test_import_preview_missing_canonical_column_reports_found_columns(client):
    import_csv = (
        "수입신고번호,원산지,HS Code,란번호,행번호,Part Number,규격,수량,수량단위\n"
        "A,CN,8708309000,004,01,MTG011114,STEERING RACK,1256,PC\n"
    )

    response = client.post(
        "/api/imports/preview",
        files={"file": ("imports.csv", import_csv, "text/csv")},
    )

    assert response.status_code == 400
    detail = response.json()["detail"]
    assert "import_accepted_date" in detail
    assert "Found columns" in detail
    assert "수입신고번호" in detail


def test_export_preview_allows_optional_description_and_unit_price(client):
    export_csv = (
        "export_date,origin,part_number,required_qty\n"
        "2025-08-16,CN,MTG011114,800\n"
    )

    response = client.post(
        "/api/exports/preview",
        files={"file": ("exports.csv", export_csv, "text/csv")},
    )

    assert response.status_code == 200
    assert response.json()["new_count"] == 1


def test_import_preview_page_shows_canonical_field_descriptions(client):
    import_csv = (
        "import_declaration_no,declaration_date,origin,hs_code,line_no,row_no,part_number,spec,import_qty,qty_unit\n"
        "A,2025-01-16,CN,8708309000,004,01,MTG011114,STEERING RACK,1256,PC\n"
    )

    response = client.post(
        "/upload/imports/preview",
        files={"file": ("imports.csv", import_csv, "text/csv")},
    )

    assert response.status_code == 200
    assert "import_accepted_date" in response.text
    assert "수입신고 수리일" in response.text
    assert "품번 / Part Number" in response.text


def test_upload_page_explains_video_style_excel_files(client):
    response = client.get("/upload")

    assert response.status_code == 200
    assert "엑셀 파일 넣기" in response.text
    assert "수입 Stock 엑셀" in response.text
    assert "수출 엑셀" in response.text
    assert "원상태진행" in response.text
    assert "수출 또는 수출 양식" in response.text
    assert "전체" not in response.text
    assert "신규" not in response.text
    assert "충돌" not in response.text


def test_upload_review_detail_delete_and_invalidate_workflow(client):
    import_csv = (
        "import_declaration_no,declaration_date,origin,hs_code,line_no,row_no,part_number,spec,import_qty,qty_unit\n"
        "A,2025-01-16,CN,8708309000,004,01,MTG011114,STEERING RACK,1256,PC\n"
    )

    preview = client.post(
        "/api/imports/preview",
        files={"file": ("imports.csv", import_csv, "text/csv")},
    )
    batch_id = preview.json()["batch_id"]

    detail = client.get(f"/upload/reviews/{batch_id}")
    assert detail.status_code == 200
    assert "행별 결과" in detail.text
    assert "인식한 컬럼 보기" in detail.text

    upload_page = client.get("/upload")
    assert "확인" in upload_page.text
    assert "삭제" in upload_page.text
    assert "배치" not in upload_page.text

    delete = client.post("/upload/delete", data={"batch_id": batch_id}, follow_redirects=False)
    assert delete.status_code == 303
    assert client.get(f"/upload/reviews/{batch_id}").status_code == 404

    second_preview = client.post(
        "/api/imports/preview",
        files={"file": ("imports.csv", import_csv, "text/csv")},
    )
    confirmed_batch_id = second_preview.json()["batch_id"]
    assert client.post("/api/imports/confirm", data={"batch_id": confirmed_batch_id}).status_code == 200

    delete_confirmed = client.post("/upload/delete", data={"batch_id": confirmed_batch_id})
    assert delete_confirmed.status_code == 400
    assert "무효 처리" in delete_confirmed.text

    invalidate = client.post(
        "/upload/invalidate",
        data={"batch_id": confirmed_batch_id, "reason": "잘못 올린 수입 파일"},
        follow_redirects=True,
    )
    assert invalidate.status_code == 200
    assert "무효 처리" in invalidate.text

    detail_after_invalidate = client.get(f"/upload/reviews/{confirmed_batch_id}")
    assert "무효 처리된 파일입니다" in detail_after_invalidate.text
    assert "잘못 올린 수입 파일" in detail_after_invalidate.text


def test_invalidated_confirmed_upload_is_excluded_from_inventory_matching_and_reports(client):
    import_csv = (
        "import_declaration_no,declaration_date,origin,hs_code,line_no,row_no,part_number,spec,import_qty,qty_unit\n"
        "A,2025-01-16,CN,8708309000,004,01,MTG011114,STEERING RACK,100,PC\n"
    )
    export_csv = "export_date,origin,part_number,required_qty\n2025-08-16,CN,MTG011114,60\n"

    import_preview = client.post("/api/imports/preview", files={"file": ("imports.csv", import_csv, "text/csv")})
    import_batch_id = import_preview.json()["batch_id"]
    assert client.post("/api/imports/confirm", data={"batch_id": import_batch_id}).status_code == 200

    export_preview = client.post("/api/exports/preview", files={"file": ("exports.csv", export_csv, "text/csv")})
    export_batch_id = export_preview.json()["batch_id"]
    assert client.post("/api/exports/confirm", data={"batch_id": export_batch_id}).status_code == 200

    assert client.post("/api/matching/run").json()["allocation_count"] == 1

    invalidate = client.post(
        "/upload/invalidate",
        data={"batch_id": import_batch_id, "reason": "중복 업로드"},
        follow_redirects=True,
    )
    assert invalidate.status_code == 200

    inventory = client.get("/api/inventory", params={"part_number": "MTG011114", "origin": "CN"})
    assert inventory.json()["total_imported_qty"] == 0
    assert inventory.json()["remaining_qty"] == 0
    assert inventory.json()["lots"] == []

    allocations = client.get("/api/reports/export-allocations").json()
    assert allocations == []

    matching = client.post("/api/matching/run")
    assert matching.json()["insufficient_stock_count"] == 1


def test_invalidated_import_upload_can_be_reactivated_by_reupload(client):
    original_import_csv = (
        "import_declaration_no,declaration_date,origin,hs_code,line_no,row_no,part_number,spec,import_qty,qty_unit\n"
        "A,2025-01-16,CN,8708309000,004,01,MTG011114,STEERING RACK,100,PC\n"
    )
    corrected_import_csv = (
        "import_declaration_no,declaration_date,origin,hs_code,line_no,row_no,part_number,spec,import_qty,qty_unit\n"
        "A,2025-01-16,CN,8708309000,004,01,MTG011114,STEERING RACK CORRECTED,120,PC\n"
    )

    preview = client.post("/api/imports/preview", files={"file": ("imports.csv", original_import_csv, "text/csv")})
    original_batch_id = preview.json()["batch_id"]
    assert client.post("/api/imports/confirm", data={"batch_id": original_batch_id}).status_code == 200

    invalidate = client.post(
        "/upload/invalidate",
        data={"batch_id": original_batch_id, "reason": "수량 수정"},
        follow_redirects=True,
    )
    assert invalidate.status_code == 200

    reupload_preview = client.post(
        "/api/imports/preview",
        files={"file": ("imports.csv", corrected_import_csv, "text/csv")},
    )
    assert reupload_preview.status_code == 200
    assert reupload_preview.json()["new_count"] == 0
    assert reupload_preview.json()["reactivate_count"] == 1

    reupload_batch_id = reupload_preview.json()["batch_id"]
    reupload_confirm = client.post("/api/imports/confirm", data={"batch_id": reupload_batch_id})
    assert reupload_confirm.status_code == 200
    assert reupload_confirm.json()["inserted_count"] == 0
    assert reupload_confirm.json()["reactivated_count"] == 1

    inventory = client.get("/api/inventory", params={"part_number": "MTG011114", "origin": "CN"})
    body = inventory.json()
    assert body["total_imported_qty"] == 120
    assert body["remaining_qty"] == 120
    assert body["lots"][0]["spec"] == "STEERING RACK CORRECTED"


def test_exports_page_shows_current_matching_rules(client):
    response = client.get("/exports")

    assert response.status_code == 200
    assert "매칭 기준 보기" in response.text
    assert "품번 동일" in response.text
    assert "원산지 동일" in response.text
    assert "720일 이내" in response.text
    assert "잔량 &gt; 0" in response.text
    assert "FIFO" in response.text
    assert "Description/단가/Amount는 매칭 판단에서 제외됩니다" in response.text


def test_reports_page_uses_clear_download_copy_and_term_explanations(client):
    response = client.get("/reports")

    assert response.status_code == 200
    assert "수출 파일 기준 결과 엑셀을 다운로드합니다." in response.text
    assert "결과 엑셀 다운로드" in response.text
    assert "매칭 결과 CSV" not in response.text
    assert "전체 리포트 XLSX" not in response.text
    assert "공모전 양식 XLSX" not in response.text
    assert "수입신고별 잔량" in response.text
    assert "수출건별 매칭 결과" in response.text
    assert "부족 수량" in response.text
    assert "Order No" in response.text
    assert "수입 재고가 없습니다" in response.text
    assert "매칭 결과가 없습니다" in response.text


def test_inventory_page_translates_status_filter_labels(client):
    response = client.get("/inventory")

    assert response.status_code == 200
    assert "조회" in response.text
    assert "사용 가능" in response.text
    assert "만료 예정" in response.text
    assert "용어 보기" in response.text
    assert "수리일" in response.text
    assert "수입신고 수리일" in response.text
